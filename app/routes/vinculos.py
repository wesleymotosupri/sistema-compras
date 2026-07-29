from fastapi import APIRouter, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from app.database import engine
import pandas as pd
import io

router = APIRouter()

FORNECEDORES_MAP = {
    "embus": "Embus", "tmac": "Tmac", "solidez": "Solidez",
    "atacado": "Atacado", "catimoto": "Catimoto", "atec": "Atec"
}

def normalizar_fornecedor(nome: str) -> str:
    """Normaliza nome do fornecedor para o padrão do sistema."""
    if not nome:
        return ""
    nome_lower = nome.lower().strip()
    # Remove acentos simples
    nome_lower = nome_lower.replace("é","e").replace("ê","e").replace("ã","a").replace("ç","c")
    for key, val in FORNECEDORES_MAP.items():
        if key in nome_lower:
            return val
    # Capitaliza primeira letra como fallback
    return nome.strip().title()

@router.get("/vinculos")
def listar_vinculos(busca: str = "", fornecedor: str = ""):
    with engine.connect() as conn:
        query = """
            SELECT v.id, v.meu_codigo, COALESCE(p.descricao, '') as descricao,
                   v.fornecedor, v.codigo_fornecedor
            FROM vinculos v
            LEFT JOIN meus_produtos p ON p.codigo = v.meu_codigo
            WHERE 1=1
        """
        params = {}
        if busca:
            query += " AND (v.meu_codigo ILIKE :busca OR p.descricao ILIKE :busca)"
            params["busca"] = f"%{busca}%"
        if fornecedor:
            query += " AND v.fornecedor = :fornecedor"
            params["fornecedor"] = fornecedor
        query += " ORDER BY v.meu_codigo, v.fornecedor LIMIT 500"
        rows = conn.execute(text(query), params).fetchall()
    return [dict(r._mapping) for r in rows]

@router.post("/vinculos")
def criar_vinculo(body: dict):
    with engine.connect() as conn:
        conn.execute(text("""
            INSERT INTO vinculos (meu_codigo, fornecedor, codigo_fornecedor)
            VALUES (:meu_codigo, :fornecedor, :codigo_fornecedor)
            ON CONFLICT DO NOTHING
        """), body)
        conn.commit()
    return {"ok": True}

@router.delete("/vinculos/{id}")
def deletar_vinculo(id: int):
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM vinculos WHERE id = :id"), {"id": id})
        conn.commit()
    return {"ok": True}

@router.get("/vinculos/exportar-excel")
def exportar_vinculos(busca: str = "", fornecedor: str = ""):
    with engine.connect() as conn:
        query = """
            SELECT v.meu_codigo, COALESCE(p.descricao, '') as descricao,
                   v.fornecedor, v.codigo_fornecedor
            FROM vinculos v
            LEFT JOIN meus_produtos p ON p.codigo = v.meu_codigo
            WHERE 1=1
        """
        params = {}
        if busca:
            query += " AND (v.meu_codigo ILIKE :busca OR p.descricao ILIKE :busca)"
            params["busca"] = f"%{busca}%"
        if fornecedor:
            query += " AND v.fornecedor = :fornecedor"
            params["fornecedor"] = fornecedor
        query += " ORDER BY v.meu_codigo, v.fornecedor"
        rows = conn.execute(text(query), params).fetchall()

    df = pd.DataFrame([dict(r._mapping) for r in rows])
    df.columns = ["Meu Código", "Descrição", "Fornecedor", "Cód. Fornecedor"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vínculos")
        ws = writer.sheets["Vínculos"]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="00A65A", end_color="00A65A", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vinculos.xlsx"}
    )

@router.get("/vinculos/modelo-excel")
def modelo_vinculos():
    """Gera planilha modelo para importação de vínculos."""
    fornecedores_ref = ["Embus", "Tmac", "Solidez", "Atacado", "Catimoto", "Atec"]
    df = pd.DataFrame({
        "Meu Codigo": ["12345", "12346"],
        "Descricao": ["DESCRIÇÃO DO PRODUTO", "OUTRO PRODUTO"],
        "Fornecedor": ["Embus", "Solidez"],
        "Codigo Fornecedor": ["16670-E", "10013"]
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vinculos")
        ws = writer.sheets["Vinculos"]
        from openpyxl.styles import PatternFill, Font, Alignment, PatternFill
        from openpyxl.styles import Border, Side
        # Cabeçalho verde
        header_fill = PatternFill(start_color="00A65A", end_color="00A65A", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        # Ajusta largura
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

        # Aba de referência com fornecedores
        ws2 = writer.book.create_sheet("Fornecedores Referencia")
        ws2["A1"] = "Fornecedores disponíveis (use exatamente estes nomes):"
        ws2["A1"].font = Font(bold=True)
        for i, f in enumerate(fornecedores_ref, start=2):
            ws2[f"A{i}"] = f
        ws2.column_dimensions["A"].width = 40
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_vinculos.xlsx"}
    )

@router.post("/vinculos/importar-excel")
async def importar_vinculos_excel(file: UploadFile = File(...)):
    """Importa vínculos em massa via Excel sem apagar vínculos existentes."""
    content = await file.read()
    df = pd.read_excel(io.BytesIO(content), dtype=str)
    df = df.fillna("")

    # Normaliza nomes das colunas
    df.columns = [c.strip().lower().replace(" ", "_").replace("ó","o").replace("ç","c") for c in df.columns]

    # Aceita variações de nome de coluna
    col_map = {}
    for col in df.columns:
        if "codigo" in col and "fornecedor" not in col:
            col_map["meu_codigo"] = col
        elif "descr" in col:
            col_map["descricao"] = col
        elif "fornecedor" in col and "codigo" not in col:
            col_map["fornecedor"] = col
        elif "codigo" in col and "fornecedor" in col:
            col_map["codigo_fornecedor"] = col

    if not all(k in col_map for k in ["meu_codigo", "fornecedor", "codigo_fornecedor"]):
        return {"erro": f"Colunas não encontradas. Encontradas: {list(df.columns)}"}

    inseridos = 0
    ignorados = 0

    with engine.connect() as conn:
        for _, row in df.iterrows():
            meu_codigo       = str(row[col_map["meu_codigo"]]).strip()
            fornecedor_raw   = str(row[col_map["fornecedor"]]).strip()
            codigo_fornecedor = str(row[col_map["codigo_fornecedor"]]).strip()

            if not meu_codigo or not codigo_fornecedor or meu_codigo == "nan":
                ignorados += 1
                continue

            fornecedor = normalizar_fornecedor(fornecedor_raw)
            if not fornecedor:
                ignorados += 1
                continue

            # Insere produto se não existe
            descricao = str(row.get(col_map.get("descricao", ""), "")).strip()
            if descricao and descricao != "nan":
                conn.execute(text("""
                    INSERT INTO meus_produtos (codigo, descricao)
                    VALUES (:codigo, :descricao)
                    ON CONFLICT (codigo) DO NOTHING
                """), {"codigo": meu_codigo, "descricao": descricao})

            # Insere vínculo — não apaga existentes, apenas adiciona novos
            result = conn.execute(text("""
                INSERT INTO vinculos (meu_codigo, fornecedor, codigo_fornecedor)
                VALUES (:meu_codigo, :fornecedor, :codigo_fornecedor)
                ON CONFLICT DO NOTHING
            """), {"meu_codigo": meu_codigo, "fornecedor": fornecedor, "codigo_fornecedor": codigo_fornecedor})

            if result.rowcount > 0:
                inseridos += 1
            else:
                ignorados += 1

        conn.commit()

    return {"inseridos": inseridos, "ignorados": ignorados}

@router.post("/vinculos/importar-csv")
async def importar_csv(file: UploadFile = File(...)):
    content = await file.read()
    try:
        df = pd.read_csv(io.BytesIO(content), dtype=str)
    except:
        df = pd.read_csv(io.BytesIO(content), dtype=str, encoding="latin1")
    df = df.fillna("")

    produtos_inseridos = 0
    vinculos_inseridos = 0
    ignorados = 0

    with engine.connect() as conn:
        for _, row in df.iterrows():
            meu_codigo        = str(row.get("meu_codigo", "")).strip()
            descricao         = str(row.get("descricao", "")).strip()
            fornecedor        = str(row.get("fornecedor", "")).strip()
            codigo_fornecedor = str(row.get("codigo_fornecedor", "")).strip()

            if not meu_codigo or not fornecedor or not codigo_fornecedor:
                ignorados += 1
                continue

            r1 = conn.execute(text("""
                INSERT INTO meus_produtos (codigo, descricao)
                VALUES (:codigo, :descricao)
                ON CONFLICT (codigo) DO NOTHING
            """), {"codigo": meu_codigo, "descricao": descricao})
            if r1.rowcount > 0:
                produtos_inseridos += 1

            r2 = conn.execute(text("""
                INSERT INTO vinculos (meu_codigo, fornecedor, codigo_fornecedor)
                VALUES (:meu_codigo, :fornecedor, :codigo_fornecedor)
                ON CONFLICT DO NOTHING
            """), {"meu_codigo": meu_codigo, "fornecedor": fornecedor, "codigo_fornecedor": codigo_fornecedor})
            if r2.rowcount > 0:
                vinculos_inseridos += 1
            else:
                ignorados += 1

        conn.commit()

    return {"produtos_inseridos": produtos_inseridos, "vinculos_inseridos": vinculos_inseridos, "ignorados": ignorados}

@router.get("/fornecedores")
def listar_fornecedores():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT DISTINCT fornecedor FROM vinculos ORDER BY fornecedor")).fetchall()
    return [r.fornecedor for r in rows]

@router.get("/produtos/busca")
def buscar_produtos(q: str = ""):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT codigo, descricao FROM meus_produtos
            WHERE codigo ILIKE :q OR descricao ILIKE :q
            LIMIT 10
        """), {"q": f"%{q}%"}).fetchall()
    return [dict(r._mapping) for r in rows]