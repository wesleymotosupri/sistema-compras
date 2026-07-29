from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from app.database import engine
import pandas as pd
import io
import datetime
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

router = APIRouter()

CNPJ_EMPRESA  = "20376718000194"
NOME_EMPRESA  = "MOTOSUPRI DISTRIBUIDORA LTDA"
Xlgr          = "MARIA COELHO AGUIAR"
NRO           = "573"
XCPL          = "CONJ B GALPAO32"
XBAIRRO       = "JARDIM SAO LUIS"
XMUN          = "SAO PAULO"
UF            = "SP"
CEP           = "05805000"
IE            = "143585224118"
FONE          = "1123482359"

@router.post("/pedidos/emitir")
def emitir_pedido(body: dict):
    fornecedor  = body.get("fornecedor", "")
    itens       = body.get("itens", [])
    empresa     = int(body.get("empresa", 1))

    if not itens:
        return {"erro": "Sem itens"}

    valor_total = sum(float(i.get("preco", 0)) * int(i.get("quantidade", 0)) for i in itens)

    with engine.connect() as conn:
        # O número do pedido vem do id do banco (SERIAL) — único e sequencial de forma atômica,
        # sem risco de corrida ou de dois pedidos saírem com o mesmo número.
        result = conn.execute(text("""
            INSERT INTO pedidos (numero, fornecedor, valor_total, criado_em)
            VALUES ('', :fornecedor, :valor_total, NOW())
            RETURNING id
        """), {"fornecedor": fornecedor, "valor_total": valor_total})
        pedido_id = result.fetchone().id
        numero = str(pedido_id).zfill(4)
        conn.execute(text("UPDATE pedidos SET numero = :numero WHERE id = :id"),
                     {"numero": numero, "id": pedido_id})
        pedido_itens_params = [
            {
                "pedido_id": pedido_id,
                "meu_codigo": item.get("meu_codigo", ""),
                "descricao": item.get("descricao", ""),
                "codigo_fornecedor": item.get("codigo_fornecedor", ""),
                "preco": float(item.get("preco", 0)),
                "quantidade": int(item.get("quantidade", 0)),
                "subtotal": float(item.get("preco", 0)) * int(item.get("quantidade", 0)),
            }
            for item in itens
        ]
        conn.execute(text("""
            INSERT INTO pedido_itens (pedido_id, meu_codigo, descricao, codigo_fornecedor, preco, quantidade, subtotal)
            VALUES (:pedido_id, :meu_codigo, :descricao, :codigo_fornecedor, :preco, :quantidade, :subtotal)
        """), pedido_itens_params)
        conn.commit()

    return {"numero": numero, "id": pedido_id, "valor_total": valor_total, "total_itens": len(itens)}

@router.get("/pedidos")
def listar_pedidos():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT id, numero, fornecedor, valor_total, criado_em FROM pedidos ORDER BY id DESC")).fetchall()
    return [dict(r._mapping) for r in rows]

@router.get("/pedidos/{pedido_id}/itens")
def itens_pedido(pedido_id: int):
    with engine.connect() as conn:
        pedido = conn.execute(text("SELECT * FROM pedidos WHERE id = :id"), {"id": pedido_id}).fetchone()
        itens  = conn.execute(text("SELECT * FROM pedido_itens WHERE pedido_id = :id ORDER BY id"), {"id": pedido_id}).fetchall()
    if not pedido:
        return {"erro": "Pedido não encontrado"}
    return {"pedido": dict(pedido._mapping), "itens": [dict(i._mapping) for i in itens]}

@router.delete("/pedidos/{pedido_id}")
def deletar_pedido(pedido_id: int):
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM pedidos WHERE id = :id"), {"id": pedido_id})
        conn.commit()
    return {"ok": True}

@router.post("/pedidos/{pedido_id}/exportar-excel")
def exportar_excel_pedido(pedido_id: int):
    with engine.connect() as conn:
        pedido = conn.execute(text("SELECT * FROM pedidos WHERE id = :id"), {"id": pedido_id}).fetchone()
        itens  = conn.execute(text("SELECT * FROM pedido_itens WHERE pedido_id = :id ORDER BY id"), {"id": pedido_id}).fetchall()
    rows = [{"Cód. Fornecedor": i.codigo_fornecedor, "Descrição": i.descricao,
             "Preço Unit.": float(i.preco), "Qtd.": int(i.quantidade), "Subtotal": float(i.subtotal)} for i in itens]
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Pedido {pedido.numero}")
        ws = writer.sheets[f"Pedido {pedido.numero}"]
        from openpyxl.styles import PatternFill, Font, Alignment
        hf = PatternFill(start_color="00A65A", end_color="00A65A", fill_type="solid")
        for cell in ws[1]:
            cell.fill = hf; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 50)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pedido_{pedido.numero}.xlsx"})

@router.post("/pedidos/{pedido_id}/exportar-pdf")
def exportar_pdf_pedido(pedido_id: int):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    with engine.connect() as conn:
        pedido = conn.execute(text("SELECT * FROM pedidos WHERE id = :id"), {"id": pedido_id}).fetchone()
        itens  = conn.execute(text("SELECT * FROM pedido_itens WHERE pedido_id = :id ORDER BY id"), {"id": pedido_id}).fetchall()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    GREEN = colors.HexColor("#00a65a"); DARK = colors.HexColor("#222d32")
    elements = []
    elements.append(Paragraph(f"Pedido de Compras #{pedido.numero}", ParagraphStyle('t', fontSize=16, textColor=DARK, fontName='Helvetica-Bold', spaceAfter=4)))
    data_fmt = pedido.criado_em.strftime("%d/%m/%Y %H:%M") if pedido.criado_em else ""
    elements.append(Paragraph(f"Fornecedor: {pedido.fornecedor}  |  Data: {data_fmt}  |  Total: R$ {float(pedido.valor_total):,.2f}",
        ParagraphStyle('s', fontSize=10, textColor=colors.HexColor("#6c757d"), spaceAfter=12)))
    elements.append(Table([[""]], colWidths=[doc.width], style=TableStyle([('LINEBELOW',(0,0),(-1,-1),1.5,GREEN)])))
    elements.append(Spacer(1, 0.4*cm))
    header = ["Cód. Fornecedor", "Descrição", "Preço Unit.", "Qtd.", "Subtotal"]
    data = [header] + [[i.codigo_fornecedor or "", i.descricao or "", f"R$ {float(i.preco):,.2f}", str(int(i.quantidade)), f"R$ {float(i.subtotal):,.2f}"] for i in itens]
    total = sum(float(i.subtotal) for i in itens)
    data.append(["", "TOTAL", "", "", f"R$ {total:,.2f}"])
    t = Table(data, colWidths=[3.5*cm, 9*cm, 2.5*cm, 1.5*cm, 2.8*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),GREEN),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),9),('ALIGN',(0,0),(-1,0),'CENTER'),('TOPPADDING',(0,0),(-1,0),6),('BOTTOMPADDING',(0,0),(-1,0),6),
        ('FONTSIZE',(0,1),(-1,-2),8),('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,colors.HexColor("#f8f9fa")]),
        ('GRID',(0,0),(-1,-2),0.3,colors.HexColor("#dee2e6")),('TOPPADDING',(0,1),(-1,-2),4),('BOTTOMPADDING',(0,1),(-1,-2),4),
        ('ALIGN',(2,1),(2,-1),'RIGHT'),('ALIGN',(3,1),(3,-1),'CENTER'),('ALIGN',(4,1),(4,-1),'RIGHT'),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('BACKGROUND',(0,-1),(-1,-1),colors.HexColor("#e8f5e9")),
        ('TEXTCOLOR',(0,-1),(-1,-1),colors.HexColor("#155724")),('LINEABOVE',(0,-1),(-1,-1),1,GREEN),
        ('TOPPADDING',(0,-1),(-1,-1),6),('BOTTOMPADDING',(0,-1),(-1,-1),6),
    ]))
    elements.append(t)
    doc.build(elements)
    output.seek(0)
    return StreamingResponse(output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=pedido_{pedido.numero}.pdf"})

def _montar_xml_bytes(identificador: int, numero_label: str, itens: list) -> bytes:
    """
    Monta o XML NF-e a partir de uma lista de itens (dicts com meu_codigo, descricao, preco, quantidade).
    Não depende do banco — pode ser usada tanto pra pedidos registrados quanto pra exportação direta.
    """
    now = datetime.datetime.now()
    dh  = now.strftime("%Y-%m-%dT%H:%M:%S") + "-03:00"
    nNF = str(identificador).zfill(10)
    cNF = str(identificador % 100000000).zfill(8)

    vProd_total = sum(float(i["preco"]) * int(i["quantidade"]) for i in itens)

    root = Element("nfeProc", versao="4.00", xmlns="http://www.portalfiscal.inf.br/nfe")
    nfe  = SubElement(root, "NFe", xmlns="http://www.portalfiscal.inf.br/nfe")
    chave = f"NFe35{now.strftime('%y%m')}{CNPJ_EMPRESA}55001{nNF}{cNF}1"
    infNFe = SubElement(nfe, "infNFe", Id=chave, versao="4.00")

    ide = SubElement(infNFe, "ide")
    for tag, val in [("cUF","35"),("cNF",cNF),("natOp","Compra para comercialização"),
                     ("mod","55"),("serie","1"),("nNF",nNF),("dhEmi",dh),("dhSaiEnt",dh),
                     ("tpNF","0"),("idDest","1"),("cMunFG","0"),("tpImp","1"),("tpEmis","1"),
                     ("cDV","3"),("tpAmb","1"),("finNFe","1"),("indFinal","0"),("indPres","9"),
                     ("procEmi","0"),("verProc","1.0.0")]:
        SubElement(ide, tag).text = val

    emit = SubElement(infNFe, "emit")
    SubElement(emit, "CNPJ").text = CNPJ_EMPRESA
    SubElement(emit, "xNome").text = NOME_EMPRESA
    end = SubElement(emit, "enderEmit")
    for tag, val in [("xLgr",Xlgr),("nro",NRO),("xCpl",XCPL),("xBairro",XBAIRRO),
                     ("cMun","0"),("xMun",XMUN),("UF",UF),("CEP",CEP),("cPais","1058"),("xPais","Brasil"),("fone",FONE)]:
        SubElement(end, tag).text = val
    SubElement(emit, "IE").text = IE
    SubElement(emit, "CRT").text = "3"

    dest = SubElement(infNFe, "dest")
    SubElement(dest, "CNPJ").text = CNPJ_EMPRESA
    SubElement(dest, "xNome").text = NOME_EMPRESA
    end2 = SubElement(dest, "enderDest")
    for tag, val in [("xLgr",Xlgr),("nro",NRO),("xBairro",XBAIRRO),("cMun","0"),
                     ("xMun",XMUN),("UF",UF),("CEP",CEP),("cPais","1058"),("xPais","Brasil")]:
        SubElement(end2, tag).text = val
    SubElement(dest, "indIEDest").text = "1"
    SubElement(dest, "IE").text = IE

    for n, item in enumerate(itens, start=1):
        qtd   = float(item["quantidade"])
        preco = float(item["preco"])
        vprod = round(preco * qtd, 2)

        det = SubElement(infNFe, "det", nItem=str(n))
        prod = SubElement(det, "prod")
        for tag, val in [("cProd", str(item["meu_codigo"])),
                         ("cEAN","SEM CÓDIGO DE BARRAS"),
                         ("xProd", str(item["descricao"])[:120]),
                         ("NCM","00000000"),("CFOP","1102"),("uCom","PC"),
                         ("qCom", f"{qtd:.4f}"),
                         ("vUnCom", f"{preco:.10f}"),
                         ("vProd", f"{vprod:.2f}"),
                         ("cEANTrib","SEM CÓDIGO DE BARRAS"),
                         ("uTrib","PC"),("qTrib", f"{qtd:.4f}"),
                         ("vUnTrib", f"{preco:.10f}"),("indTot","1")]:
            SubElement(prod, tag).text = val

        imposto = SubElement(det, "imposto")
        icms = SubElement(imposto, "ICMS")
        icms40 = SubElement(icms, "ICMS40")
        SubElement(icms40, "orig").text = "0"
        SubElement(icms40, "CST").text = "40"

        ipi = SubElement(imposto, "IPI")
        SubElement(ipi, "cEnq").text = "999"
        ipitrib = SubElement(ipi, "IPITrib")
        for tag, val in [("CST","99"),("vBC","0.00"),("pIPI","0.00"),("vIPI","0.00")]:
            SubElement(ipitrib, tag).text = val

        pis = SubElement(imposto, "PIS")
        pisoutr = SubElement(pis, "PISOutr")
        for tag, val in [("CST","49"),("vBC","0.00"),("pPIS","0.00"),("vPIS","0.00")]:
            SubElement(pisoutr, tag).text = val

        cofins = SubElement(imposto, "COFINS")
        cofinsoutr = SubElement(cofins, "COFINSOutr")
        for tag, val in [("CST","49"),("vBC","0.00"),("pCOFINS","0.00"),("vCOFINS","0.00")]:
            SubElement(cofinsoutr, tag).text = val

        SubElement(det, "vItem").text = f"{vprod:.2f}"

    total_el = SubElement(infNFe, "total")
    icmstot = SubElement(total_el, "ICMSTot")
    vp = round(vProd_total, 2)
    for tag, val in [("vBC",f"{vp:.2f}"),("vICMS","0.00"),("vICMSDeson","0.00"),("vFCP","0.00"),
                     ("vBCST","0.00"),("vST","0.00"),("vFCPST","0.00"),("vFCPSTRet","0.00"),
                     ("vProd",f"{vp:.2f}"),("vFrete","0.00"),("vSeg","0.00"),("vDesc","0.00"),
                     ("vII","0.00"),("vIPI","0.00"),("vIPIDevol","0.00"),("vPIS","0.00"),
                     ("vCOFINS","0.00"),("vOutro","0.00"),("vNF",f"{vp:.2f}")]:
        SubElement(icmstot, tag).text = val

    transp = SubElement(infNFe, "transp")
    SubElement(transp, "modFrete").text = "9"

    pag = SubElement(infNFe, "pag")
    detPag = SubElement(pag, "detPag")
    SubElement(detPag, "tPag").text = "90"
    SubElement(detPag, "vPag").text = "0.00"

    infAdic = SubElement(infNFe, "infAdic")
    SubElement(infAdic, "infCpl").text = f"PEDIDO LOJA 2 - #{numero_label}"

    prot = SubElement(root, "protNFe", versao="4.00")
    infProt = SubElement(prot, "infProt")
    for tag, val in [("tpAmb","1"),("verAplic","SVRS"),("chNFe",chave.replace("NFe","")),
                     ("dhRecbto",dh),("nProt","135260520376718"),("digVal","MzUyNjA1MjAzNzY3MTgwMDAxOTQ1"),
                     ("cStat","100"),("xMotivo","Autorizado o uso da NF-e")]:
        SubElement(infProt, tag).text = val

    xml_str = tostring(root, encoding="unicode")
    return minidom.parseString(xml_str).toprettyxml(indent="  ", encoding="UTF-8")


@router.post("/pedido/exportar-xml")
async def exportar_xml_direto(body: dict):
    """
    Gera o XML direto a partir dos itens já carregados na tela (igual o Exportar Excel) —
    NÃO cria pedido no banco, então é uma chamada só, rápida.
    """
    itens = body.get("itens", [])
    if not itens:
        return {"erro": "Sem itens"}
    itens_norm = [
        {
            "meu_codigo": i.get("meu_codigo") or i.get("codprod", ""),
            "descricao": i.get("descricao") or i.get("descrprod", ""),
            "preco": float(i.get("preco") or i.get("custo_gerencial") or 0),
            "quantidade": int(i.get("quantidade") or i.get("sugestao_compra_ajustada") or 0),
        }
        for i in itens
    ]
    itens_norm = [i for i in itens_norm if i["quantidade"] > 0]
    if not itens_norm:
        return {"erro": "Nenhum item com quantidade"}

    identificador = int(datetime.datetime.now().timestamp())
    xml_bytes = _montar_xml_bytes(identificador, "PREVIA", itens_norm)

    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=pedido.xml"}
    )


@router.post("/pedidos/{pedido_id}/exportar-xml")
def exportar_xml_pedido(pedido_id: int):
    """Gera XML no formato NF-e para importação no Sankhya — Pedido Filial (a partir de um pedido já registrado)."""
    with engine.connect() as conn:
        pedido = conn.execute(text("SELECT * FROM pedidos WHERE id = :id"), {"id": pedido_id}).fetchone()
        itens  = conn.execute(text("SELECT * FROM pedido_itens WHERE pedido_id = :id ORDER BY id"), {"id": pedido_id}).fetchall()

    itens_norm = [
        {"meu_codigo": i.meu_codigo, "descricao": i.descricao, "preco": i.preco, "quantidade": i.quantidade}
        for i in itens
    ]
    xml_bytes = _montar_xml_bytes(pedido_id, pedido.numero, itens_norm)

    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename=pedido_{pedido.numero}.xml"}
    )
    xml_str = tostring(root, encoding="unicode")
    xml_pretty = minidom.parseString(xml_str).toprettyxml(indent="  ", encoding="UTF-8")

    return StreamingResponse(
        io.BytesIO(xml_pretty),
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename=pedido_{pedido.numero}.xml"}
    )