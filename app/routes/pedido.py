from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from app.database import engine
from app.sankhya import buscar_lista_compras, buscar_marcas
import pandas as pd
import io

router = APIRouter()

FORNECEDORES = ["Embus", "Tmac", "Solidez", "Atacado", "Catimoto", "Atec"]

def _cruzar_precos(produtos: list[dict]) -> list[dict]:
    if not produtos:
        return []

    codigos = [str(p["codprod"]) for p in produtos]

    with engine.connect() as conn:
        vinculos = conn.execute(text("""
            SELECT v.meu_codigo, v.fornecedor, v.codigo_fornecedor,
                   COALESCE(p.preco, 0) AS preco
            FROM vinculos v
            LEFT JOIN precos p
              ON p.fornecedor = v.fornecedor
             AND p.codigo_fornecedor = v.codigo_fornecedor
        """)).fetchall()
        codigos_set = set(codigos)
        vinculos = [v for v in vinculos if str(v.meu_codigo) in codigos_set]

    mapa = {}
    for v in vinculos:
        codigo = str(v.meu_codigo)
        if codigo not in mapa:
            mapa[codigo] = {}
        mapa[codigo][v.fornecedor] = {
            "preco": float(v.preco) if v.preco else None,
            "codigo_fornecedor": v.codigo_fornecedor
        }

    resultado = []
    for p in produtos:
        codigo = str(p["codprod"])
        vinculos_prod = mapa.get(codigo, {})

        precos_forn = {}
        for forn in FORNECEDORES:
            info = vinculos_prod.get(forn)
            precos_forn[forn] = {
                "preco": info["preco"] if info and info["preco"] else None,
                "codigo_fornecedor": info["codigo_fornecedor"] if info else None
            }

        disponiveis = {f: v for f, v in precos_forn.items() if v["preco"]}
        melhor = min(disponiveis, key=lambda f: disponiveis[f]["preco"]) if disponiveis else None

        resultado.append({
            **p,
            "codigo": codigo,
            "descricao": p.get("descrprod", ""),
            "precos": precos_forn,
            "fornecedor_sugerido": melhor,
            "preco_sugerido": disponiveis[melhor]["preco"] if melhor else None,
            "cod_forn_sugerido": disponiveis[melhor]["codigo_fornecedor"] if melhor else None,
            "tem_vinculo": bool(disponiveis)  # True só se tem vínculo COM preço
        })

    return resultado


@router.get("/pedido/apoio-compras")
def apoio_compras():
    from app.sankhya import buscar_apoio_compras
    return buscar_apoio_compras()


@router.get("/pedido/marcas")
def marcas():
    return buscar_marcas()


@router.post("/pedido/sankhya")
def pedido_do_sankhya(body: dict):
    # ---- Parâmetros que vêm da tela ----
    codemp            = int(body.get("codemp", 1))
    codlocal          = int(body.get("codlocal", 10000000))
    dias_uteis_mes    = int(body.get("dias_uteis_mes", 22))
    cobertura_dias    = int(body.get("cobertura_dias", 30))
    tipos_venda       = str(body.get("tipos_venda", "1100"))
    margem_seguranca  = float(body.get("margem_seguranca", 1.2))
    qtd_sem_historico = int(body.get("qtd_sem_historico", 3))
    tipos_compra      = str(body.get("tipos_compra", "109,110,205,215"))

    incluir_apoio = [v.strip() for v in body.get("incluir_apoio", "").split(",") if v.strip()]
    excluir_apoio = [v.strip() for v in body.get("excluir_apoio", "").split(",") if v.strip()]
    incluir_marca = [v.strip() for v in body.get("incluir_marca", "").split(",") if v.strip()]
    excluir_marca = [v.strip() for v in body.get("excluir_marca", "").split(",") if v.strip()]

    produtos = buscar_lista_compras(
        codemp=codemp,
        codlocal=codlocal,
        dias_uteis_mes=dias_uteis_mes,
        cobertura_dias=cobertura_dias,
        tipos_venda=tipos_venda,
        margem_seguranca=margem_seguranca,
        qtd_sem_historico=qtd_sem_historico,
        tipos_compra=tipos_compra,
        incluir_apoio=incluir_apoio,
        excluir_apoio=excluir_apoio,
        incluir_marca=incluir_marca,
        excluir_marca=excluir_marca,
    )

    # Compra externa (fornecedor) só faz sentido pra Empresa 1.
    # Empresa 3/4 é transferência interna da matriz -> usa custo_gerencial, não cruza fornecedor.
    if codemp == 1:
        return _cruzar_precos(produtos)
    else:
        return produtos


@router.post("/pedido/exportar")
async def exportar_pedido(body: dict):
    itens  = body.get("itens", [])
    codemp = int(body.get("codemp", 1))
    rows   = []

    if codemp == 1:
        for item in itens:
            rows.append({
                "Código":          item.get("codigo", ""),
                "Descrição":       item.get("descricao", ""),
                "Fornecedor":      item.get("fornecedor_selecionado", ""),
                "Cód. Fornecedor": item.get("cod_forn_selecionado", ""),
                "Preço":           item.get("preco_selecionado", ""),
                "Qtd.":            item.get("sugestao_compra_ajustada", ""),
                "Subtotal":        item.get("subtotal", ""),
            })
    else:
        for item in itens:
            rows.append({
                "Código":          item.get("codprod", ""),
                "Descrição":       item.get("descrprod", ""),
                "Marca":           item.get("marca", ""),
                "Custo Gerencial": item.get("custo_gerencial", ""),
                "Qtd.":            item.get("sugestao_compra_ajustada", ""),
            })

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pedido")
        ws = writer.sheets["Pedido"]
        from openpyxl.styles import PatternFill, Font, Alignment
        hf = PatternFill(start_color="00A65A", end_color="00A65A", fill_type="solid")
        for cell in ws[1]:
            cell.fill = hf; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 50)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pedido.xlsx"}
    )


@router.post("/pedido/exportar-sem-vinculo")
async def exportar_sem_vinculo(body: dict):
    itens = body.get("itens", [])
    df = pd.DataFrame(itens)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sem Vinculo")
        ws = writer.sheets["Sem Vinculo"]
        from openpyxl.styles import PatternFill, Font, Alignment
        hf = PatternFill(start_color="DD4B39", end_color="DD4B39", fill_type="solid")
        for cell in ws[1]:
            cell.fill = hf; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 50)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sem_vinculo.xlsx"}
    )